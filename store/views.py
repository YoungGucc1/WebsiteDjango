from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Prefetch
from django.utils.text import slugify # Added this import
from .models import Product, Category, Tag, Image, Price, Warehouse, Stock, Counterparty, WorkerProductAudit
from .forms import ProductForm, ImageForm, CategoryForm, TagForm, PriceForm, WarehouseForm, StockForm, CounterpartyForm, WorkerProductAuditForm, ExcelUploadForm, ManualImageUploadForm # Added ManualImageUploadForm
from django.http import HttpResponse
from openpyxl import load_workbook
from decimal import Decimal
import os
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
import io
import base64
from django.core.files.base import ContentFile
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

# Basic Views
def home(request):
    products = Product.objects.filter(is_active=True).prefetch_related(
        Prefetch('images', queryset=Image.objects.filter(is_main=True), to_attr='main_product_image'),
        Prefetch('prices', queryset=Price.objects.filter(is_active=True, price_type=Price.PriceType.SELLING).order_by('amount'), to_attr='selling_prices')
    )

    # Efficiently get the main image and lowest selling price
    for product in products:
        product.display_image = product.main_product_image[0] if product.main_product_image else (product.images.first() if product.images.exists() else None)
        product.display_price = product.selling_prices[0] if product.selling_prices else None
        
    categories = Category.objects.filter(parent__isnull=True).prefetch_related('children') # Top-level categories
    return render(request, 'store/home.html', {'products': products, 'categories': categories})

def product_list(request, category_slug=None):
    category = None
    products_query = Product.objects.filter(is_active=True).select_related('category').prefetch_related(
        Prefetch('images', queryset=Image.objects.filter(is_main=True), to_attr='main_product_image_list'),
        Prefetch('prices', queryset=Price.objects.filter(is_active=True, price_type=Price.PriceType.SELLING).order_by('amount'), to_attr='selling_prices_list')
    )

    if category_slug:
        category = get_object_or_404(Category, slug=category_slug)
        products_query = products_query.filter(category=category)
    
    products = list(products_query) # Evaluate the queryset

    for product in products:
        product.display_image = product.main_product_image_list[0] if product.main_product_image_list else (product.images.first() if product.images.exists() else None)
        product.display_price = product.selling_prices_list[0] if product.selling_prices_list else None

    categories = Category.objects.all()
    return render(request, 'store/product_list.html', {'products': products, 'category': category, 'categories': categories})

def product_detail(request, slug):
    product = get_object_or_404(Product.objects.prefetch_related(
        'images', 
        Prefetch('prices', queryset=Price.objects.filter(is_active=True).order_by('price_type', 'amount'))
    ), slug=slug, is_active=True)
    
    main_image = product.images.filter(is_main=True).first()
    if not main_image:
        main_image = product.images.first()
        
    other_images = product.images.exclude(id=main_image.id) if main_image else product.images.all()
    
    # Get the first active selling price, or any active price if no selling price
    current_price = product.prices.filter(price_type=Price.PriceType.SELLING, is_active=True).order_by('amount').first()
    if not current_price:
        current_price = product.prices.filter(is_active=True).order_by('amount').first()

    return render(request, 'store/product_detail.html', {
        'product': product,
        'main_image': main_image,
        'other_images': other_images,
        'current_price': current_price
    })

def category_list(request):
    categories = Category.objects.filter(parent__isnull=True).prefetch_related('children__children') # Fetch up to 3 levels
    return render(request, 'store/category_list.html', {'categories': categories})

# Worker Product Audit Views
@login_required
def worker_product_list(request):
    # Fetch all products and their audit status
    # Using select_related for product and prefetch_related for audit details
    # to optimize database queries.
    products_with_audit = []
    all_products = Product.objects.filter(is_active=True).order_by('name')

    for product in all_products:
        audit_details, created = WorkerProductAudit.objects.get_or_create(
            product=product,
            defaults={'last_audited_by': None} # Default if creating
        )
        # If created, it means no audit existed, so it's pending.
        # If get_or_create found an existing one, its is_completed status is used.
        products_with_audit.append({
            'product': product,
            'audit_details': audit_details,
            'is_completed': audit_details.is_completed
        })

    return render(request, 'store/worker_product_list.html', {'products_with_audit': products_with_audit})

@login_required
def worker_product_audit_view(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    # Get or create an audit instance for this product
    audit_instance, created = WorkerProductAudit.objects.get_or_create(
        product=product,
        defaults={'last_audited_by': request.user if request.user.is_authenticated else None}
    )

    if request.method == 'POST':
        form = WorkerProductAuditForm(request.POST, request.FILES, instance=audit_instance)
        if form.is_valid():
            audit_to_save = form.save(commit=False)
            audit_to_save.last_audited_by = request.user # Ensure current user is set
            
            # The form's save method handles creating/updating the Image instance
            # and linking it to photo_taken.
            # The WorkerProductAudit model's save method handles setting is_completed.
            audit_to_save.save()
            
            return redirect('store:worker_product_list')
    else:
        form = WorkerProductAuditForm(instance=audit_instance)

    return render(request, 'store/worker_product_audit_form.html', {
        'form': form,
        'product': product,
        'audit_instance': audit_instance
    })


# --- Excel Import View ---
@login_required # Ensure only logged-in users can access
def import_products_from_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return HttpResponse("Invalid file format. Please upload an Excel file.", status=400)

            try:
                workbook = load_workbook(excel_file)
                sheet = workbook.active # Assumes data is in the first sheet

                # --- Header Mapping (Adjust based on your Excel file) ---
                # Example: {'Excel Column Name': 'model_field_name'}
                header_map = {
                    'Наименование': 'name',
                    'Наименование (доп)': 'name_2',
                    'Артикул': 'article_number',
                    'Описание': 'description',
                    'Краткое описание': 'short_description',
                    'Категория': 'category_name', # Will be used to find/create Category object
                    'Теги': 'tags_str',           # Comma-separated string of tags
                    'Цена': 'selling_price',
                    'Валюта продажи': 'selling_currency',
                    'Закуп': 'purchase_price',
                    'Валюта закупки': 'purchase_currency',
                    'Остаток': 'stock_quantity',
                    'Склад': 'warehouse_name',
                    'Изображение (путь или URL)': 'image_path_or_url', # Handle local path or URL
                    'Главное изображение': 'is_main_image', # 'Да' or 'Нет'
                    'Активен': 'is_active_str', # 'Да' or 'Нет'
                    'В избранном': 'is_featured_str', # 'Да' or 'Нет'
                }
                
                headers = [cell.value for cell in sheet[1]] # Get headers from the first row
                
                # Validate required headers
                required_excel_headers = ['Наименование', 'Категория', 'Цена'] # Example
                for req_header in required_excel_headers:
                    if req_header not in headers:
                        return HttpResponse(f"Missing required Excel column: {req_header}", status=400)

                products_created_count = 0
                products_updated_count = 0
                errors = []

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row): # Skip empty rows
                        continue

                    product_data = {}
                    for col_idx, cell_value in enumerate(row):
                        header_name = headers[col_idx]
                        model_field = header_map.get(header_name)
                        if model_field:
                            product_data[model_field] = cell_value
                    
                    if not product_data.get('name'):
                        errors.append(f"Row {row_idx}: Product name is missing. Skipping.")
                        continue

                    try:
                        # --- Handle Category ---
                        category_name = product_data.pop('category_name', None)
                        if not category_name:
                            errors.append(f"Row {row_idx} (Product: {product_data.get('name')}): Category name is missing. Skipping.")
                            continue
                        category, _ = Category.objects.get_or_create(name=category_name, defaults={'slug': slugify(category_name)})

                        # --- Handle Tags ---
                        tags_str = product_data.pop('tags_str', '')
                        tag_instances = []
                        if tags_str:
                            tag_names = [name.strip() for name in tags_str.split(',') if name.strip()]
                            for tag_name in tag_names:
                                tag, _ = Tag.objects.get_or_create(name=tag_name, defaults={'slug': slugify(tag_name)})
                                tag_instances.append(tag)
                        
                        # --- Handle Prices ---
                        selling_price_val = product_data.pop('selling_price', None)
                        selling_currency_str = product_data.pop('selling_currency', 'KZT') # Default currency
                        purchase_price_val = product_data.pop('purchase_price', None)
                        purchase_currency_str = product_data.pop('purchase_currency', 'KZT')

                        # --- Handle Stock ---
                        stock_quantity_val = product_data.pop('stock_quantity', 0)
                        warehouse_name_str = product_data.pop('warehouse_name', 'Основной склад') # Default warehouse
                        warehouse, _ = Warehouse.objects.get_or_create(name=warehouse_name_str, defaults={'slug': slugify(warehouse_name_str)})

                        # --- Handle Boolean Fields ---
                        is_active = str(product_data.pop('is_active_str', 'Да')).lower() in ['да', 'yes', 'true', '1']
                        is_featured = str(product_data.pop('is_featured_str', 'Нет')).lower() in ['да', 'yes', 'true', '1']

                        # --- Prepare Product Fields ---
                        # Remove price/stock related fields that are not direct Product model fields
                        product_fields = {k: v for k, v in product_data.items() if hasattr(Product, k) and k not in ['image_path_or_url', 'is_main_image']}
                        product_fields['category'] = category
                        product_fields['is_active'] = is_active
                        product_fields['is_featured'] = is_featured
                        
                        # --- Create or Update Product ---
                        # Use article_number as a unique identifier if available, otherwise name
                        identifier_field = 'article_number' if product_fields.get('article_number') else 'name'
                        identifier_value = product_fields.get(identifier_field)

                        if identifier_field == 'name' and not identifier_value: # Should have been caught earlier
                            errors.append(f"Row {row_idx}: Product name is missing after processing. Skipping.")
                            continue
                        
                        product_instance = None
                        if identifier_value:
                            try:
                                if identifier_field == 'article_number':
                                    product_instance = Product.objects.get(article_number=identifier_value)
                                else: # name
                                    product_instance = Product.objects.get(name=identifier_value, category=category) # Name might not be unique across categories
                                
                                # Update existing product
                                for key, value in product_fields.items():
                                    setattr(product_instance, key, value)
                                product_instance.save()
                                products_updated_count += 1
                            except Product.DoesNotExist:
                                pass # Will create new below

                        if not product_instance:
                            product_instance = Product.objects.create(**product_fields)
                            products_created_count += 1
                        
                        # --- Set Tags (after product is saved) ---
                        if tag_instances:
                            product_instance.tags.set(tag_instances)

                        # --- Handle Prices (after product is saved) ---
                        if selling_price_val is not None:
                            try:
                                selling_price_decimal = Decimal(str(selling_price_val))
                                Price.objects.update_or_create(
                                    product=product_instance,
                                    price_type=Price.PriceType.SELLING,
                                    currency=selling_currency_str.upper(),
                                    defaults={'amount': selling_price_decimal, 'is_active': True}
                                )
                            except ValueError:
                                errors.append(f"Row {row_idx} (Product: {product_instance.name}): Invalid selling price format '{selling_price_val}'.")
                        
                        if purchase_price_val is not None:
                            try:
                                purchase_price_decimal = Decimal(str(purchase_price_val))
                                Price.objects.update_or_create(
                                    product=product_instance,
                                    price_type=Price.PriceType.PURCHASE,
                                    currency=purchase_currency_str.upper(),
                                    defaults={'amount': purchase_price_decimal, 'is_active': True}
                                )
                            except ValueError:
                                errors.append(f"Row {row_idx} (Product: {product_instance.name}): Invalid purchase price format '{purchase_price_val}'.")

                        # --- Handle Stock (after product is saved) ---
                        if stock_quantity_val is not None:
                            try:
                                stock_quantity_int = int(stock_quantity_val)
                                Stock.objects.update_or_create(
                                    product=product_instance,
                                    warehouse=warehouse,
                                    defaults={'quantity': stock_quantity_int}
                                )
                            except ValueError:
                                 errors.append(f"Row {row_idx} (Product: {product_instance.name}): Invalid stock quantity format '{stock_quantity_val}'.")


                        # --- Handle Image ---
                        image_path_or_url = product_data.get('image_path_or_url')
                        is_main_image_str = str(product_data.get('is_main_image', 'Нет')).lower()
                        is_main = is_main_image_str in ['да', 'yes', 'true', '1']

                        if image_path_or_url:
                            try:
                                img_name_default = f"{slugify(product_instance.name)}_{product_instance.id.hex[:8]}"
                                
                                # Check if it's a local file path
                                # This is a basic check; more robust path validation might be needed
                                if os.path.exists(image_path_or_url) and os.path.isfile(image_path_or_url):
                                    with open(image_path_or_url, 'rb') as f:
                                        # Create an InMemoryUploadedFile for Django's ImageField
                                        img_file = InMemoryUploadedFile(
                                            file=io.BytesIO(f.read()),
                                            field_name='file_path',
                                            name=os.path.basename(image_path_or_url),
                                            content_type='image/jpeg', # Adjust content type if needed
                                            size=os.path.getsize(image_path_or_url),
                                            charset=None
                                        )
                                    image_instance = Image.objects.create(
                                        name=img_name_default,
                                        file_path=img_file,
                                        type=Image.ImageType.PRODUCT,
                                        is_main=is_main 
                                    )
                                    product_instance.images.add(image_instance)
                                # TODO: Add handling for image URLs (download and save)
                                # else: (handle URL)
                                #    errors.append(f"Row {row_idx} (Product: {product_instance.name}): Image path '{image_path_or_url}' not found or not a file. URL import not yet implemented.")
                            except Exception as e:
                                errors.append(f"Row {row_idx} (Product: {product_instance.name}): Error processing image '{image_path_or_url}': {e}")
                    
                    except Exception as e:
                        errors.append(f"Row {row_idx} (Product: {product_data.get('name', 'Unknown')}): General error - {e}")
                
                # --- Prepare Summary ---
                summary_message = f"Import completed. Products created: {products_created_count}, Products updated: {products_updated_count}."
                if errors:
                    summary_message += "\nErrors encountered:\n" + "\n".join(errors)
                
                return HttpResponse(summary_message, content_type="text/plain")

            except Exception as e:
                return HttpResponse(f"An error occurred during processing: {e}", status=500)
        else:
            # If form is not valid on POST, re-render with errors
            return render(request, 'store/import_products.html', {'form': form})
    else:
        form = ExcelUploadForm() # Instantiate a new form for GET requests
    
    return render(request, 'store/import_products.html', {'form': form})

# --- Auto Image Search View ---
from .image_search_utils import get_image_search_results, save_image_for_product
from .forms import ImageSearchSettingsForm # Import the new form

@login_required # Or use @staff_member_required for admin-only access
def auto_search_product_images_view(request):
    """
    A view to allow users to set parameters and trigger automatic image
    searching and then select images to save.
    """
    context = {'title': "Automatic Product Image Search"}
    form = ImageSearchSettingsForm(request.POST or None)
    context['form'] = form

    if request.method == 'POST' and form.is_valid():
        num_results = form.cleaned_data['num_results']
        img_size = form.cleaned_data['img_size']
        img_type = form.cleaned_data['img_type']
        img_color_type = form.cleaned_data['img_color_type']
        file_type = form.cleaned_data['file_type']
        safe_search = form.cleaned_data['safe_search']

        products_with_images = []
        products_with_errors = []
        all_products = Product.objects.filter(is_active=True)

        for product in all_products:
            try:
                image_urls = get_image_search_results(
                    product,
                    num_results=num_results,
                    img_size=img_size,
                    img_type=img_type,
                    img_color_type=img_color_type,
                    file_type=file_type,
                    safe_search=safe_search
                )
                if image_urls:
                    products_with_images.append({
                        'product': product,
                        'image_urls': image_urls
                    })
            except Exception as e:
                error_message = f"Error processing product {product.name} (ID: {product.id}): {e}"
                print(error_message)
                products_with_errors.append({'name': product.name, 'id': product.id, 'error': str(e)})

        context.update({
            'products_with_images': products_with_images,
            'products_with_errors': products_with_errors,
            'search_triggered': True
        })
        return render(request, 'store/auto_search_results.html', context)

    return render(request, 'store/auto_image_search_form.html', context)

@login_required
def save_selected_images_view(request):
    if request.method == 'POST':
        product_id_next = request.POST.get('next_product_id')
        skip_product = request.POST.get('skip_product') # Check for skip button

        if skip_product and product_id_next:
            # If 'Skip' button was pressed, go to next product without saving
            return redirect('store:one_by_one_image_search', product_id=product_id_next)
        
        selected_images = request.POST.getlist('selected_images')
        
        for item in selected_images:
            product_id, image_url = item.split(',', 1)
            product = get_object_or_404(Product, id=product_id)
            is_main = f'is_main_{product_id}' in request.POST and request.POST[f'is_main_{product_id}'] == image_url
            save_image_for_product(product, image_url, is_main=is_main)

        if product_id_next:
            return redirect('store:one_by_one_image_search', product_id=product_id_next)
        else:
            return redirect('store:one_by_one_product_list')

    return redirect('store:auto_search_product_images')

@login_required
def upload_webcam_image(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        image_data = request.POST.get('image_data')
        
        if image_data:
            try:
                # Decode the base64 image data
                format, imgstr = image_data.split(';base64,') 
                ext = format.split('/')[-1] 
                
                # Generate a filename
                timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                image_index = product.images.count()
                filename = f"webcam_{slugify(product.name)}_{image_index + 1}_{timestamp}.{ext}"

                # Create a Django ContentFile
                data = ContentFile(base64.b64decode(imgstr), name=filename)
                
                # Use the related manager's create() method to create, save, and associate the image in one step.
                # This is safer and avoids potential double-saving issues.
                store_image = product.images.create(
                    name=f"{product.name} Webcam Image {image_index + 1}",
                    alt_text=f"Webcam image for {product.name}",
                    type=Image.ImageType.PRODUCT,
                    is_main=False,
                    file_path=data
                )
                
                return JsonResponse({'success': True, 'image_url': store_image.file_path.url})
            except Exception as e:
                print(f"Error saving webcam image: {e}")
                return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
def upload_product_image(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ManualImageUploadForm(request.POST, request.FILES)
        if form.is_valid():
            image_file = form.cleaned_data['image_file']
            # Use the existing save_image_for_product utility
            save_image_for_product(product, image_file) # Pass file object directly
            # Optionally, add a success message
            # messages.success(request, "Image uploaded successfully!")
        else:
            # Optionally, add an error message
            # messages.error(request, "Error uploading image.")
            pass # Handle form errors if needed, for now just pass
    return redirect('store:one_by_one_image_search', product_id=product.id)


# --- One by One Image Search Views ---

@login_required
def one_by_one_product_list(request):
    products = Product.objects.filter(is_active=True).order_by('name').prefetch_related('images')
    for product in products:
        product.has_images = product.images.exists()
    return render(request, 'store/one_by_one_product_list.html', {'products': products})

@login_required
def one_by_one_image_search(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    context = {'product': product}

    # Initialize forms
    search_form = ImageSearchSettingsForm(request.POST if request.method == 'POST' else None)
    upload_form = ManualImageUploadForm() # Always instantiate for GET requests

    if request.method == 'POST':
        # Check if the POST is for search settings or image upload
        if 'num_results' in request.POST: # This is a search settings submission
            form = search_form
            if form.is_valid():
                num_results = form.cleaned_data['num_results']
                img_size = form.cleaned_data['img_size']
                img_type = form.cleaned_data['img_type']
                img_color_type = form.cleaned_data['img_color_type']
                file_type = form.cleaned_data['file_type']
                safe_search = form.cleaned_data['safe_search']

                image_urls = get_image_search_results(
                    product,
                    num_results=num_results,
                    img_size=img_size,
                    img_type=img_type,
                    img_color_type=img_color_type,
                    file_type=file_type,
                    safe_search=safe_search
                )
                context['image_urls'] = image_urls
                context['search_triggered'] = True
            else:
                context['search_form'] = form # Pass invalid form back
                context['upload_form'] = upload_form
                return render(request, 'store/one_by_one_image_search.html', context)
        # No need for an 'else' for upload form here, as upload is handled by a separate view
    else: # GET request
        # Automatically trigger search on GET if it's a fresh load for a product
        # This handles redirection from 'save_selected_images_view' or direct access
        form = search_form # Use the initialized search_form for defaults
        num_results = form.initial.get('num_results', 7)
        img_size = form.initial.get('img_size', 'large')
        img_type = form.initial.get('img_type', 'photo')
        img_color_type = form.initial.get('img_color_type', 'any')
        file_type = form.initial.get('file_type', 'jpg')
        safe_search = form.initial.get('safe_search', 'active')

        image_urls = get_image_search_results(
            product,
            num_results=num_results,
            img_size=img_size,
            img_type=img_type,
            img_color_type=img_color_type,
            file_type=file_type,
            safe_search=safe_search
        )
        context['image_urls'] = image_urls
        context['search_triggered'] = True

    context['search_form'] = search_form # Always pass search form
    context['upload_form'] = upload_form # Always pass upload form

    # For the "Save and Next" button (logic remains the same)
    next_product = Product.objects.filter(is_active=True, name__gt=product.name).order_by('name').first()
    if next_product:
        context['next_product_id'] = next_product.id

    return render(request, 'store/one_by_one_image_search.html', context)

from django.urls import reverse

def metro_home_view(request):
    # Fetch top-level categories and their direct children
    categories = Category.objects.filter(parent__isnull=True).prefetch_related('children')

    # Fetch products (e.g., latest 20 active products)
    # Reusing the prefetch logic from the 'home' view for consistency
    products_query = Product.objects.filter(is_active=True).order_by('-created_at').prefetch_related(
        Prefetch('images', queryset=Image.objects.filter(is_main=True), to_attr='main_product_image_list'), # Use a different to_attr if home uses main_product_image
        Prefetch('prices', queryset=Price.objects.filter(is_active=True, price_type=Price.PriceType.SELLING).order_by('amount'), to_attr='selling_prices_list')
    )[:24] # Show more products

    products = list(products_query)

    for product in products:
        # Prepare display image
        if product.main_product_image_list:
            product.display_image = product.main_product_image_list[0]
        elif product.images.exists():
            product.display_image = product.images.first()
        else:
            product.display_image = None # Template will handle placeholder

        # Prepare display price
        if product.selling_prices_list:
            product.display_price = product.selling_prices_list[0]
        else: # Fallback to any active price if no selling price
            any_active_price = product.prices.filter(is_active=True).order_by('amount').first()
            product.display_price = any_active_price
            
        # Add a default URL for the product card link
        try:
            product.default_url = reverse('store:product_detail', kwargs={'slug': product.slug})
        except Exception: # Handle cases where reverse might fail (e.g. no slug)
            product.default_url = "#"


    # Add default URLs for categories
    for category in categories:
        try:
            category.default_url = reverse('store:product_list_by_category', kwargs={'category_slug': category.slug})
        except Exception:
            category.default_url = "#"
        if hasattr(category, 'children'):
            for child in category.children.all():
                try:
                    child.default_url = reverse('store:product_list_by_category', kwargs={'category_slug': child.slug})
                except Exception:
                    child.default_url = "#"
                    
    # The template uses category.get_absolute_url_if_exists_else_default_url
    # and product.get_absolute_url_if_exists_else_default_url
    # For now, I'll pass the objects and let the template try to access 'default_url'
    # or a get_absolute_url method if it exists on the model.
    # To be robust, these should be methods on the models.
    # For now, the template will use `product.default_url` and `category.default_url`
    # I will modify the template to use these.

    context = {
        'categories': categories,
        'products': products,
        'store_name': getattr(settings, 'STORE_NAME', 'Metro Store') # Example: get store name from settings
    }
    return render(request, 'store/metro_home.html', context)

def ajax_product_list(request, category_slug=None):
    products_query = Product.objects.filter(is_active=True).prefetch_related(
        Prefetch('images', queryset=Image.objects.filter(is_main=True), to_attr='main_product_image'),
        Prefetch('prices', queryset=Price.objects.filter(is_active=True, price_type=Price.PriceType.SELLING).order_by('amount'), to_attr='selling_prices')
    )

    if category_slug:
        category = get_object_or_404(Category, slug=category_slug)
        products_query = products_query.filter(category=category)
    
    products = list(products_query)

    for product in products:
        product.display_image = product.main_product_image[0] if product.main_product_image else (product.images.first() if product.images.exists() else None)
        product.display_price = product.selling_prices[0] if product.selling_prices else None
        
    return render(request, 'store/product_list_partial.html', {'products': products})
